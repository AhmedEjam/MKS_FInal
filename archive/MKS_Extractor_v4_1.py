#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MKS Smart Architecture Extractor — v4.1
========================================
What changed from v4.0
─────────────────────────────────────────────────────────────────
MULTI-FOLDER SUPPORT
  CONFIG.scan_folders lets you pick exactly which top-level
  modules to include. Default is ["app", "core", "features"].
  Missing folders are reported and skipped gracefully.
  Each row now carries a "Module" column (app / core / features)
  so you can filter or sort the sheet by module.

NON-CODER COLUMNS  (AI fills these in during the review pass)
  "Main Role"    — 2-4 words: what kind of thing this code is.
                   e.g.  Data Validator · Screen Builder · Quiz Importer
  "What It Does" — one plain-English sentence, no jargon.
                   e.g.  "Saves a completed quiz result to local storage
                          and marks the session as finished."
  Both columns sit near the left so a non-technical reader can
  scan the sheet without scrolling past code.

Dependencies
─────────────────────────────────────────────────────────────────
    pip install pandas openpyxl tqdm
    (tqdm is optional — the script runs without it)
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
    _HAS_TQDM = True
except ImportError:
    _HAS_TQDM = False
    def tqdm(it, **_):   # type: ignore[no-redef]
        return it


# ══════════════════════════════════════════════════════════════════
# 1.  CONFIGURATION
# ══════════════════════════════════════════════════════════════════

@dataclass
class Config:
    # ── Root of your Android project ─────────────────────────────
    project_root: str = "/Users/ahmedy.ajam/Android MKS"

    # ── Which top-level modules to scan ──────────────────────────
    # Each name is a direct subfolder of project_root.
    # • Use ["app", "core", "features"] to scan those three only.
    # • Use [] (empty list) to scan the entire project_root tree.
    scan_folders: list[str] = field(
        default_factory=lambda: ["app", "core", "feature"]
    )

    output_file:     str = "MKS_Smart_Architecture.xlsx"
    max_cell_chars:  int = 28_000   # Emergency cut — avoids Excel cell limit
    min_chunk_chars: int = 1_500    # Minimum content before a natural split fires
    encodings: list[str] = field(
        default_factory=lambda: ["utf-8", "utf-8-sig", "latin-1"]
    )


CONFIG = Config()


# ══════════════════════════════════════════════════════════════════
# 2.  STRING & COMMENT-AWARE BRACE COUNTING
# ══════════════════════════════════════════════════════════════════

def count_real_braces(line: str) -> tuple[int, int]:
    """
    Count *code-level* { and } only, skipping:
      • Everything after //  (line comment)
      • Content inside "…"   (string literals)
    Returns (open_count, close_count).
    """
    open_c = close_c = 0
    in_string = False
    i, n = 0, len(line)

    while i < n:
        ch = line[i]
        if not in_string and ch == "/" and i + 1 < n and line[i + 1] == "/":
            break
        if ch == '"' and (i == 0 or line[i - 1] != "\\"):
            in_string = not in_string
        if not in_string:
            if   ch == "{": open_c  += 1
            elif ch == "}": close_c += 1
        i += 1

    return open_c, close_c


# ══════════════════════════════════════════════════════════════════
# 3.  DECLARATION & ANNOTATION DETECTION
# ══════════════════════════════════════════════════════════════════

_DECL_RE = re.compile(
    r"^(?:"
    r"(?:(?:public|private|protected|internal|"
    r"override|abstract|open|final|sealed|"
    r"suspend|inline|infix|operator|tailrec|"
    r"external|expect|actual|lateinit|const|"
    r"data|inner|value|crossinline|noinline|"
    r"reified|vararg|enum)\s+)*"
    r"(?:fun\b|class\b|interface\b|object\b|typealias\b)"
    r"|companion\s+object"
    r")"
)

_ANN_RE = re.compile(r"^@\w+")


def _is_decl(s: str) -> bool:
    return bool(_DECL_RE.match(s))

def _is_ann(s: str) -> bool:
    return bool(_ANN_RE.match(s))


# ══════════════════════════════════════════════════════════════════
# 4.  ENTITY NAME & COMPONENT TYPE HEURISTICS
# ══════════════════════════════════════════════════════════════════

_ENTITY_RE = re.compile(
    r"\b(?:fun|class|interface|object|typealias)\s+(\w+)"
)

def extract_entity_name(chunk: str) -> str:
    first = "\n".join(chunk.splitlines()[:12])
    m = _ENTITY_RE.search(first)
    return m.group(1) if m else ""


_TYPE_PATTERNS: list[tuple[str, str]] = [
    ("DAO",                  r"@Dao\b|@Query\b|@Insert\b|@Update\b|@Delete\b|@Upsert\b|@Transaction\b"),
    ("Entity / Schema",      r"@Entity\b|@PrimaryKey\b|@ColumnInfo\b|@ForeignKey\b|@Embedded\b"),
    ("DI — Hilt",            r"@HiltViewModel\b|@Module\b|@InstallIn\b|@Provides\b|@Binds\b|@Inject\b"),
    ("Migration",            r"\bMigration\s*\(|addMigrations\b|database\.execSQL\b"),
    ("UI (Composable)",      r"@Composable\b|Scaffold\b|LazyColumn\b|LazyRow\b|TextField\s*\(|Button\s*\("),
    ("ViewModel",            r"\bViewModel\s*\(\)|MutableStateFlow\b|StateFlow\b|viewModelScope\b"),
    ("Repository",           r"\bRepository\b|withContext\s*\(Dispatchers|\.collect\s*\{"),
    ("Navigation",           r"NavController\b|NavGraph\b|NavHost\b|composable\s*\(\s*route|BackHandler\b"),
    ("Data Model",           r"\bdata class\b|\bsealed class\b|@Serializable\b|@Parcelize\b"),
    ("Utility / Extension",  r"object \w+(?:Helper|Utils|Extensions|Ext)\b"),
]

def infer_component_type(chunk: str) -> str:
    for label, pattern in _TYPE_PATTERNS:
        if re.search(pattern, chunk):
            return label
    return ""


# ══════════════════════════════════════════════════════════════════
# 5.  SMART CHUNKER
# ══════════════════════════════════════════════════════════════════

def smart_split_kotlin_code(content: str) -> list[dict]:
    """
    Split Kotlin source into logical chunks.
    Each chunk → { code, start_line, end_line }

    Rules (in priority order):
      1. Emergency cut  — next line would overflow max_cell_chars.
      2. Natural split  — at a declaration/annotation boundary when
                          prev brace level ≤ 1 and chunk ≥ min_chunk_chars.
      3. Annotation guard — never split between @Annotation and its fun/class.
    """
    chunks:      list[dict] = []
    cur_lines:   list[str]  = []
    cur_len      = 0
    brace_level  = 0
    in_ann_block = False
    chunk_start  = 1

    for lineno, line in enumerate(content.splitlines(), start=1):
        s        = line.strip()
        line_len = len(line) + 1

        # Brace level: snapshot BEFORE this line
        prev_level = brace_level
        opens, closes = count_real_braces(s)
        brace_level = max(0, brace_level + opens - closes)

        is_decl = _is_decl(s)
        is_ann  = _is_ann(s)

        is_new_block  = (is_ann or is_decl) and not in_ann_block
        is_safe_split = (prev_level <= 1) and is_new_block

        # ── Emergency cut ────────────────────────────────────────
        if cur_len + line_len > CONFIG.max_cell_chars:
            cur_lines.append("// ✂ [TBC — continued in next chunk]")
            if cur_lines:
                chunks.append(_chunk(cur_lines, chunk_start, lineno - 1))
            cur_lines    = ["// ✂ [Continued from previous chunk]"]
            cur_len      = len(cur_lines[0]) + 1
            chunk_start  = lineno
            in_ann_block = False

        # ── Natural split ────────────────────────────────────────
        elif is_safe_split and cur_len > CONFIG.min_chunk_chars:
            if cur_lines:
                chunks.append(_chunk(cur_lines, chunk_start, lineno - 1))
            cur_lines    = []
            cur_len      = 0
            chunk_start  = lineno
            in_ann_block = False

        # ── Annotation-block state machine ───────────────────────
        if is_ann:
            in_ann_block = True
        elif is_decl or (s and not s.startswith("//")):
            in_ann_block = False

        cur_lines.append(line)
        cur_len += line_len

    if cur_lines:
        tail = _chunk(cur_lines, chunk_start, len(content.splitlines()))
        if tail["code"].strip():
            chunks.append(tail)

    return chunks


def _chunk(lines: list[str], start: int, end: int) -> dict:
    return {"code": "\n".join(lines).strip(), "start_line": start, "end_line": end}


# ══════════════════════════════════════════════════════════════════
# 6.  RESILIENT FILE READER
# ══════════════════════════════════════════════════════════════════

def read_file(path: Path) -> Optional[str]:
    for enc in CONFIG.encodings:
        try:
            return path.read_text(encoding=enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return None


# ══════════════════════════════════════════════════════════════════
# 7.  EXCEL SCHEMA & FORMATTING
# ══════════════════════════════════════════════════════════════════

# Column order defines sheet column order.
# Width is in Excel character units.
COLUMNS: dict[str, int] = {
    # ── Identity ─────────────────────────────────────────────────
    "ID":                       14,
    "Module":                   12,   # app / core / features
    "File Name":                26,
    "Entity Name":              22,

    # ── Non-coder insight (AI fills these) ───────────────────────
    "Main Role":                22,   # 2-4 words  e.g. "Data Validator"
    "What It Does":             58,   # 1 sentence e.g. "Saves quiz result to local DB"

    # ── Position in file ─────────────────────────────────────────
    "Start Line":               11,
    "End Line":                 11,

    # ── Code & technical analysis (AI fills these) ───────────────
    "Code Snippet":             80,
    "Auto-detected Type":       22,
    "Component Type":           22,
    "AI Analysis & Summary":    55,
    "Refactoring Suggestions":  45,
    "Dependencies":             35,

    # ── Navigation ───────────────────────────────────────────────
    "File Path":                48,
    "Status":                   12,
}

# Column index of the code cell (1-based, derived from COLUMNS order)
_CODE_COL_IDX = list(COLUMNS.keys()).index("Code Snippet") + 1

# Non-coder columns get a distinct header colour so they stand out
_NONCODER_COLS = {"Main Role", "What It Does"}
_NONCODER_COL_INDICES = {
    list(COLUMNS.keys()).index(c) + 1 for c in _NONCODER_COLS
}

_HDR_FONT       = Font(name="Calibri",  bold=True, color="FFFFFF", size=11)
_HDR_FILL       = PatternFill("solid", fgColor="2C3E50")          # dark slate
_HDR_FILL_NC    = PatternFill("solid", fgColor="1A6B45")          # deep green
_HDR_FONT_NC    = Font(name="Calibri",  bold=True, color="FFFFFF", size=11)
_COD_FONT       = Font(name="Consolas", size=9)
_BOD_FONT       = Font(name="Calibri",  size=10)
_NC_FONT        = Font(name="Calibri",  size=10, italic=True, color="1A4731")


def apply_excel_formatting(wb) -> None:
    ws = wb.active
    ws.title = "MKS Architecture"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws.row_dimensions[1].height = 30

    # Header row
    for idx, cell in enumerate(ws[1], start=1):
        is_nc = (idx in _NONCODER_COL_INDICES)
        cell.font      = _HDR_FONT_NC if is_nc else _HDR_FONT
        cell.fill      = _HDR_FILL_NC if is_nc else _HDR_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Column widths
    for idx, width in enumerate(COLUMNS.values(), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Body cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            is_code = (cell.column == _CODE_COL_IDX)
            is_nc   = (cell.column in _NONCODER_COL_INDICES)
            cell.font      = _COD_FONT if is_code else (_NC_FONT if is_nc else _BOD_FONT)
            cell.alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )


# ══════════════════════════════════════════════════════════════════
# 8.  MAIN
# ══════════════════════════════════════════════════════════════════

def _collect_kt_files(root: Path) -> list[Path]:
    """
    Gather .kt files from the configured scan_folders.
    Falls back to scanning the entire root if scan_folders is empty.
    """
    if not CONFIG.scan_folders:
        found = sorted(root.rglob("*.kt"))
        print(f"  📂  (full project root)  →  {len(found)} .kt files")
        return found

    print("  Scanning selected modules:\n")
    kt_files: list[Path] = []

    for folder_name in CONFIG.scan_folders:
        folder_path = root / folder_name
        if not folder_path.exists():
            print(f"     ⚠️   {folder_name:<16}  not found — skipped")
            continue
        found = sorted(folder_path.rglob("*.kt"))
        kt_files.extend(found)
        print(f"     ✓   {folder_name:<16}  {len(found):>4} .kt files")

    return sorted(set(kt_files))   # deduplicate & sort


def _derive_module(relative: Path) -> str:
    """
    Return the top-level folder name from a path relative to project_root.
    e.g.  app/src/main/java/…   →  "app"
          features/quiz/src/…   →  "features/quiz"   (two levels for features)
    """
    parts = relative.parts
    if not parts:
        return ""
    top = parts[0]
    # For multi-feature projects, show the feature sub-module as well
    if top == "features" and len(parts) >= 2:
        return f"{top}/{parts[1]}"
    return top


def main() -> None:
    root = Path(CONFIG.project_root)

    if not root.exists():
        print(f"❌  Project root not found:\n    {root}")
        sys.exit(1)

    print(f"\n🗂   Project root: {root}\n")
    kt_files = _collect_kt_files(root)

    if not kt_files:
        print("\n⚠️  No .kt files found — check CONFIG.project_root / scan_folders.")
        sys.exit(1)

    print(f"\n🔍  Total: {len(kt_files)} Kotlin files — starting extraction...\n")

    all_rows: list[dict] = []
    failed:   list[str]  = []
    file_counter = 1

    iterator = (
        tqdm(kt_files, desc="Extracting", unit="file")
        if _HAS_TQDM else kt_files
    )

    for file_path in iterator:
        relative = file_path.relative_to(root)
        content  = read_file(file_path)

        if content is None:
            failed.append(str(relative))
            if not _HAS_TQDM:
                print(f"  ⚠️  Encoding error — skipped: {relative}")
            continue

        if not content.strip():
            continue

        chunks = smart_split_kotlin_code(content)
        if not chunks:
            continue

        module = _derive_module(relative)

        for idx, c in enumerate(chunks, start=1):
            code = c["code"]
            all_rows.append({
                "ID":                       f"MKS-{file_counter:04d}-{idx:02d}",
                "Module":                   module,
                "File Name":                file_path.name,
                "Entity Name":              extract_entity_name(code),
                "Main Role":                "",   # AI: 2-4 words, e.g. "Data Validator"
                "What It Does":             "",   # AI: 1 plain-English sentence
                "Start Line":               c["start_line"],
                "End Line":                 c["end_line"],
                "Code Snippet":             code,
                "Auto-detected Type":       infer_component_type(code),
                "Component Type":           "",
                "AI Analysis & Summary":    "",
                "Refactoring Suggestions":  "",
                "Dependencies":             "",
                "File Path":                str(relative),
                "Status":                   "Pending",
            })

        file_counter += 1   # Only increments on a successfully read file

    if not all_rows:
        print("⚠️  Nothing extracted — check project root and folder names.")
        sys.exit(1)

    # ── Write Excel ───────────────────────────────────────────────
    print(f"\n💾  Writing {len(all_rows)} chunks to: {CONFIG.output_file}")
    df = pd.DataFrame(all_rows, columns=list(COLUMNS.keys()))
    df.to_excel(CONFIG.output_file, index=False, engine="openpyxl")

    print("🎨  Applying formatting...")
    wb = load_workbook(CONFIG.output_file)
    apply_excel_formatting(wb)
    wb.save(CONFIG.output_file)

    # ── Summary ───────────────────────────────────────────────────
    files_ok = file_counter - 1
    avg      = len(all_rows) / files_ok if files_ok else 0
    sep      = "─" * 54

    print(f"\n{sep}")
    print(f"  ✅  Extraction complete!")
    print(f"  📁  Files processed :  {files_ok}")
    print(f"  🧩  Chunks produced :  {len(all_rows)}")
    print(f"  📊  Avg per file    :  {avg:.1f}")
    if failed:
        print(f"  ⚠️   Failed files   :  {len(failed)}")
        for f in failed:
            print(f"       • {f}")
    print(f"  📄  Saved to        :  {CONFIG.output_file}")
    print(sep)


if __name__ == "__main__":
    main()
